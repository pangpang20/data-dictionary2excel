import pandas as pd
import dmPython
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse


# 达梦数据库连接配置
DM_DEFAULT_PORT = 5236
DM_DEFAULT_USER = "SYSDBA"


def generate_data_dictionary(db_name, user, password, host, port):
    # Create dmPython connection
    conn_string = f"{user}@{host}:{port}"
    conn = dmPython.connect(
        user=user,
        password=password,
        host=host,
        port=port,
        dbname=db_name
    )
    cursor = conn.cursor()

    # Get all table names and comments from user's schema
    table_query = """
    SELECT
        TABLE_NAME,
        COMMENTS AS TABLE_COMMENT
    FROM USER_TABLES t
    LEFT JOIN USER_TAB_COMMENTS c ON t.TABLE_NAME = c.TABLE_NAME AND c.TABLE_TYPE = 'TABLE'
    """
    cursor.execute(table_query)
    tables_df = pd.DataFrame(cursor.fetchall(), columns=['table_name', 'table_comment'])

    # Initialize DataFrame with initial rows
    data = pd.DataFrame([
        ["1.1.", f"数据库：{db_name}", "", "", "", "", ""],
        ["", f"列出的数据库对象：{len(tables_df)} 表", "", "", "", "", ""],
        ["", "", "", "", "", "", ""]  # Empty row
    ], columns=["编号", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"])

    # Query each table's structure
    for table_idx, table_row in tables_df.iterrows():
        table_name = table_row['table_name']
        table_comment = table_row['table_comment'] if table_row['table_comment'] else ""

        # Add table header and field header
        new_rows = pd.DataFrame([
            [f"", f"{table_idx + 1}.", f"表：{table_name}", "", "", "", f"表注释：{table_comment}"],
            ["", "字段", "", "", "", "", ""]
        ], columns=["编号", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"])

        # Query table structure - get column info from USER_TAB_COLUMNS
        column_query = """
        SELECT
            c.COLUMN_NAME AS "字段名称",
            c.DATA_TYPE AS "类型",
            c.DATA_LENGTH AS "长度",
            c.DATA_PRECISION AS "精度",
            c.DATA_SCALE AS "小数位",
            c.NULLABLE AS "是否允许为空",
            CASE
                WHEN pk.COLUMN_NAME IS NOT NULL THEN 'YES'
                ELSE 'NO'
            END AS "是否主键",
            cm.COMMENTS AS "中文注释"
        FROM
            USER_TAB_COLUMNS c
        LEFT JOIN USER_COL_COMMENTS cm ON c.TABLE_NAME = cm.TABLE_NAME AND c.COLUMN_NAME = cm.COLUMN_NAME
        LEFT JOIN (
            SELECT a.COLUMN_NAME
            FROM USER_CONST_COLUMNS a
            JOIN USER_CONSTRAINTS b ON a.CONSTRAINT_NAME = b.CONSTRAINT_NAME
            WHERE a.TABLE_NAME = %s AND b.CONSTRAINT_TYPE = 'P'
        ) pk ON c.COLUMN_NAME = pk.COLUMN_NAME
        WHERE
            c.TABLE_NAME = %s
        ORDER BY
            c.COLUMN_ID
        """
        cursor.execute(column_query, (table_name, table_name))
        df = pd.DataFrame(
            cursor.fetchall(),
            columns=["字段名称", "类型", "长度", "精度", "小数位", "是否允许为空", "是否主键", "中文注释"]
        )
        print(f"正在处理表：{table_name}")

        # Add position column and combine type with length/precision
        df['序号'] = range(1, len(df) + 1)
        df['类型'] = df.apply(
            lambda row: f"{row['类型']}({int(row['长度'])})"
            if pd.notnull(row['长度']) and row['类型'] not in ['TEXT', 'LONGTEXT', 'MEDIUMTEXT', 'BLOB', 'TINYBLOB', 'MEDIUMBLOB', 'LONGBLOB', 'CLOB', 'BFILE']
            else (f"{row['类型']}({int(row['精度'])},{int(row['小数位'])})"
                  if pd.notnull(row['精度']) and pd.notnull(row['小数位'])
                  else row['类型']),
            axis=1
        )
        df = df.drop(columns=['长度', '精度', '小数位'])

        # Transform 是否允许为空 from 'Y'/'N' to '是'/'否'
        df['是否允许为空'] = df['是否允许为空'].apply(lambda x: '是' if x == 'Y' else '否')

        # Reorder and reconstruct df with exact column order
        columns_order = ["编号", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"]
        df = df[["序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"]].copy()
        df.insert(0, "编号", "")
        df = df[columns_order]

        # Add field headers and data
        field_headers = pd.DataFrame([["", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"]],
                                   columns=columns_order)
        new_rows = pd.concat([new_rows, field_headers, df], ignore_index=True)
        new_rows = pd.concat([new_rows, pd.DataFrame([["", "", "", "", "", "", ""]],
                                                    columns=columns_order)],
                            ignore_index=True)

        # Append new rows to data with correct column order
        data = pd.concat([data, new_rows], ignore_index=True)

    # Close cursor and connection
    cursor.close()
    conn.close()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "数据字典"

    # Write DataFrame to worksheet without header row
    for row in data.itertuples(index=False):
        ws.append(list(row))

    # Apply styling
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Track rows for borders and bold fonts
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7), 1):
        row_data = [cell.value for cell in row]
        if row_data[0] and row_data[0].endswith('.') or row_data[1] == "序号":
            for cell in row:
                cell.font = bold_font
        if row_data[1] == "序号" or (row_data[0] == "" and isinstance(row_data[1], (int, float))):
            for cell_idx, cell in enumerate(row, 1):
                if cell_idx > 1:
                    cell.border = border

    # Save to Excel
    output_file = f"{db_name}_data_dictionary.xlsx"
    wb.save(output_file)

    print(f"数据字典已生成到 {output_file}")

def main():
    parser = argparse.ArgumentParser(description="生成达梦数据库数据字典.")
    parser.add_argument("db_name", help="要生成字典的数据库名称")
    parser.add_argument("--user", default=DM_DEFAULT_USER, help="数据库用户名")
    parser.add_argument("--password", default="", help="数据库密码")
    parser.add_argument("--host", default="", help="数据库主机")
    parser.add_argument("--port", default=str(DM_DEFAULT_PORT), help="数据库端口")
    args = parser.parse_args()

    generate_data_dictionary(args.db_name, args.user, args.password, args.host, args.port)

if __name__ == "__main__":
    main()
