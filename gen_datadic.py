import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse

def generate_data_dictionary(db_name):
    # Create SQLAlchemy engine with escaped password
    # update the username, password, and database host in the connection string below
    engine = create_engine(f'mysql+mysqlconnector://#username#:#password#@#dbhost#/{db_name}')

    # Get all table names and comments
    table_query = """
    SELECT TABLE_NAME, TABLE_COMMENT
    FROM INFORMATION_SCHEMA.TABLES
    WHERE TABLE_SCHEMA = %s;
    """
    tables_df = pd.read_sql(table_query, engine, params=(db_name,))

    # Initialize DataFrame with initial rows
    data = pd.DataFrame([
        ["1.1.", f"数据库: {db_name}", "", "", "", "", ""],
        ["", f"列出的数据库对象: {len(tables_df)} 表", "", "", "", "", ""],
        ["", "", "", "", "", "", ""]  # Empty row
    ], columns=["编号", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"])

    # Query each table's structure
    for table_idx, table_row in tables_df.iterrows():
        table_name = table_row['TABLE_NAME']
        table_comment = table_row['TABLE_COMMENT'] if table_row['TABLE_COMMENT'] else ""

        # Add table header and field header
        new_rows = pd.DataFrame([
            [f"", f"{table_idx + 1}.", f"表: {table_name}", "", "", "", f"表注释: {table_comment}"],
            ["", "字段", "", "", "", "", ""]
        ], columns=["编号", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"])

        # Query table structure
        column_query = """
        SELECT 
            COLUMN_NAME AS '字段名称',
            DATA_TYPE AS '类型',
            CHARACTER_MAXIMUM_LENGTH AS '长度',
            NUMERIC_PRECISION AS '精度',
            NUMERIC_SCALE AS '小数位',
            IS_NULLABLE AS '是否允许为空',
            COLUMN_KEY AS '是否主键',
            COLUMN_COMMENT AS '中文注释'
        FROM 
            INFORMATION_SCHEMA.COLUMNS 
        WHERE 
            TABLE_NAME = %s 
            AND TABLE_SCHEMA = %s;
        """
        df = pd.read_sql(column_query, engine, params=(table_name, db_name))
        print(f"正在处理表：{table_name}")
        # Add position column and combine type with length/precision, resetting position per table
        df['序号'] = range(1, len(df) + 1)
        df['类型'] = df.apply(
            lambda row: f"{row['类型']}({int(row['长度'])})" 
            if pd.notnull(row['长度']) and row['类型'] not in ['text', 'longtext', 'mediumtext', 'blob', 'tinyblob', 'mediumblob', 'longblob'] 
            else (f"{row['类型']}({int(row['精度'])},{int(row['小数位'])})" 
                  if row['类型'] == 'decimal' and pd.notnull(row['精度']) and pd.notnull(row['小数位']) 
                  else row['类型']),
            axis=1
        )
        df = df.drop(columns=['长度', '精度', '小数位'])

        # Transform COLUMN_KEY to 是否主键 (YES/NO)
        df['是否主键'] = df['是否主键'].apply(lambda x: 'YES' if x == 'PRI' else 'NO')

        # Reorder and reconstruct df with exact column order
        columns_order = ["编号", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"]
        df = df[["序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"]].copy()  # Select required columns
        df.insert(0, "编号", "")  # Add '编号' as the first column
        df = df[columns_order]  # Ensure exact column order

        # Add field headers and data
        field_headers = pd.DataFrame([["", "序号", "字段名称", "类型", "是否允许为空", "是否主键", "中文注释"]], 
                                   columns=columns_order)
        new_rows = pd.concat([new_rows, field_headers, df], ignore_index=True)
        new_rows = pd.concat([new_rows, pd.DataFrame([["", "", "", "", "", "", ""]], 
                                                    columns=columns_order)], 
                            ignore_index=True)

        # Append new rows to data with correct column order
        data = pd.concat([data, new_rows], ignore_index=True)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "数据字典"

    # Write DataFrame to worksheet without header row
    for row in data.itertuples(index=False):
        ws.append(list(row))

    # Apply styling
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Track rows for borders (field tables, excluding first column) and bold fonts (table/field headers)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7), 1):
        row_data = [cell.value for cell in row]
        # Bold table headers (e.g., "1. 表: table_name") and field headers (e.g., "序号,字段名称,...")
        if row_data[0] and row_data[0].endswith('.') or row_data[1] == "序号":
            for cell in row:
                cell.font = bold_font
        # Apply borders to field table rows (headers and data), excluding first column
        if row_data[1] == "序号" or (row_data[0] == "" and isinstance(row_data[1], (int, float))):
            for cell_idx, cell in enumerate(row, 1):
                if cell_idx > 1:  # Skip the first column (编号)
                    cell.border = border

    # Save to Excel
    output_file = f"{db_name}_data_dictionary.xlsx"
    wb.save(output_file)

    print(f"数据字典已生成到 {output_file}")

def main():
    parser = argparse.ArgumentParser(description="生成数据字典.")
    parser.add_argument("db_name", help="要生成字典的数据库名称")
    args = parser.parse_args()

    generate_data_dictionary(args.db_name)

if __name__ == "__main__":
    main()